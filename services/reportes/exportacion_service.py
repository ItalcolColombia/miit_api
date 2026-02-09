import io
from datetime import datetime
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer
)


class ExportacionService:
    """
    Servicio para exportar datos de reportes a diferentes formatos.
    """

    # Configuración de estilos
    HEADER_COLOR = '366092'
    HEADER_FONT_COLOR = 'FFFFFF'
    ALTERNATE_ROW_COLOR = 'F2F2F2'
    TOTALS_COLOR = 'D9E1F2'

    # Campos que deben formatearse como enteros en exportación (sin decimales)
    # Estos campos se formatearán como enteros independientemente del tipo_dato definido
    CAMPOS_ENTEROS = {
        # Nombres reales de campos (según BD)
        'transaccion_id',
        'almacen_id',
        'bascula_id',
        'consecutivo_pesada',
        # Variantes alternativas
        'id_transaccion',
        'id_almacen',
        'bascula',
        'id transaccion',
        'id transacción',
        'id almacen',
        'id almacén',
        'báscula',
        'idtransaccion',
        'idalmacen',
        'consecutivo',
    }

    # ========================================================
    # EXPORTACIÓN A EXCEL
    # ========================================================

    def exportar_excel(
            self,
            datos: List[Dict[str, Any]],
            nombre_reporte: str,
            columnas: List[Dict[str, Any]],
            totales: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Exporta datos a formato Excel (.xlsx).

        Args:
            datos: Lista de registros a exportar
            nombre_reporte: Nombre del reporte para el título
            columnas: Definición de columnas
            totales: Diccionario con totales (opcional)

        Returns:
            Contenido del archivo Excel en bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = nombre_reporte[:31]  # Límite de Excel para nombre de hoja

        # Estilos
        header_fill = PatternFill(
            start_color=self.HEADER_COLOR,
            end_color=self.HEADER_COLOR,
            fill_type="solid"
        )
        header_font = Font(color=self.HEADER_FONT_COLOR, bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        alternate_fill = PatternFill(
            start_color=self.ALTERNATE_ROW_COLOR,
            end_color=self.ALTERNATE_ROW_COLOR,
            fill_type="solid"
        )

        totals_fill = PatternFill(
            start_color=self.TOTALS_COLOR,
            end_color=self.TOTALS_COLOR,
            fill_type="solid"
        )
        totals_font = Font(bold=True, size=11)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título del reporte
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
        title_cell = ws.cell(row=1, column=1, value=nombre_reporte)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Fecha de generación
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columnas))
        date_cell = ws.cell(
            row=2,
            column=1,
            value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = Alignment(horizontal="center")

        # Fila vacía
        start_row = 4

        # Encabezados
        for col_idx, columna in enumerate(columnas, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=columna['nombre_mostrar'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        for row_idx, fila in enumerate(datos, start_row + 1):
            for col_idx, columna in enumerate(columnas, 1):
                valor = fila.get(columna['campo'], '')

                # Formatear valor según tipo
                valor_formateado = self._formatear_valor_excel(
                    valor,
                    columna.get('tipo_dato', 'string'),
                    columna.get('decimales', 2),
                    columna.get('campo', '')
                )

                cell = ws.cell(row=row_idx, column=col_idx, value=valor_formateado)
                cell.border = border

                # Alineación
                alineacion = columna.get('alineacion', 'left')
                cell.alignment = Alignment(horizontal=alineacion)

                # Filas alternadas
                if (row_idx - start_row) % 2 == 0:
                    cell.fill = alternate_fill

        # Fila de totales
        if totales:
            total_row = start_row + len(datos) + 1

            # Primera celda con "TOTALES"
            total_label_cell = ws.cell(row=total_row, column=1, value="TOTALES")
            total_label_cell.fill = totals_fill
            total_label_cell.font = totals_font
            total_label_cell.border = border

            for col_idx, columna in enumerate(columnas, 1):
                campo = columna['campo']
                if campo in totales:
                    valor = totales[campo]
                    valor_formateado = self._formatear_valor_excel(
                        valor,
                        columna.get('tipo_dato', 'number'),
                        columna.get('decimales', 2),
                        campo
                    )
                    cell = ws.cell(row=total_row, column=col_idx, value=valor_formateado)
                else:
                    cell = ws.cell(row=total_row, column=col_idx, value="")

                cell.fill = totals_fill
                cell.font = totals_font
                cell.border = border

                alineacion = columna.get('alineacion', 'left')
                cell.alignment = Alignment(horizontal=alineacion)

        # Ajustar ancho de columnas
        for col_idx, columna in enumerate(columnas, 1):
            # Calcular ancho basado en contenido
            max_length = len(str(columna['nombre_mostrar']))

            for fila in datos[:100]:  # Solo revisar primeras 100 filas
                valor = fila.get(columna['campo'], '')
                if valor:
                    max_length = max(max_length, len(str(valor)))

            # Limitar ancho
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

        # Congelar encabezados
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

        # Guardar en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file.getvalue()

    def _formatear_valor_excel(
            self,
            valor: Any,
            tipo_dato: str,
            decimales: int = 2,
            campo: str = ''
    ) -> Any:
        """
        Formatea un valor para Excel según su tipo.

        Args:
            valor: Valor a formatear
            tipo_dato: Tipo de dato de la columna
            decimales: Número de decimales para números
            campo: Nombre del campo (para detectar campos enteros)
        """
        if valor is None:
            return ""

        # Verificar si es un campo que debe ser entero
        campo_lower = campo.lower() if campo else ''
        es_campo_entero = campo_lower in self.CAMPOS_ENTEROS or tipo_dato == 'integer'

        if es_campo_entero:
            try:
                return int(float(valor))
            except (ValueError, TypeError):
                return valor

        if tipo_dato == 'number':
            try:
                return round(float(valor), decimales)
            except (ValueError, TypeError):
                return valor

        elif tipo_dato in ('date', 'datetime'):
            if isinstance(valor, datetime):
                return valor
            elif isinstance(valor, str):
                try:
                    return datetime.fromisoformat(valor.replace('Z', '+00:00'))
                except:
                    return valor

        return valor

    # ========================================================
    # EXPORTACIÓN A PDF
    # ========================================================

    def exportar_pdf(
            self,
            datos: List[Dict[str, Any]],
            nombre_reporte: str,
            columnas: List[Dict[str, Any]],
            totales: Optional[Dict[str, Any]] = None,
            orientacion: str = 'auto'
    ) -> bytes:
        """
        Exporta datos a formato PDF optimizado para reportes con muchas columnas.
        Usa Paragraph en las celdas para permitir word wrap automático.

        Args:
            datos: Lista de registros a exportar
            nombre_reporte: Nombre del reporte
            columnas: Definición de columnas
            totales: Diccionario con totales (opcional)
            orientacion: 'portrait', 'landscape' o 'auto' (detecta automáticamente)

        Returns:
            Contenido del archivo PDF en bytes
        """
        buffer = io.BytesIO()

        num_columnas = len(columnas)

        # Determinar orientación según número de columnas si es 'auto'
        # Portrait para <= 5 columnas, Landscape para > 5 columnas
        if orientacion == 'auto':
            if num_columnas > 5:
                orientacion = 'landscape'
            else:
                orientacion = 'portrait'

        # Configurar página
        if orientacion == 'landscape':
            pagesize = landscape(A4)
        else:
            pagesize = A4

        # Ajustar márgenes según número de columnas
        if num_columnas > 10:
            margins = 0.5 * cm
        elif num_columnas > 7:
            margins = 0.75 * cm
        else:
            margins = 1 * cm

        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=margins,
            leftMargin=margins,
            topMargin=1 * cm,
            bottomMargin=1 * cm
        )

        elements = []
        styles = getSampleStyleSheet()

        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14 if num_columnas <= 8 else 12,
            alignment=TA_CENTER,
            spaceAfter=8
        )

        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=TA_CENTER,
            textColor=colors.gray,
            spaceAfter=12
        )

        # Determinar tamaño de fuente según número de columnas
        if num_columnas > 12:
            font_size_header = 6
            font_size_body = 5
        elif num_columnas > 10:
            font_size_header = 7
            font_size_body = 6
        elif num_columnas > 8:
            font_size_header = 8
            font_size_body = 7
        elif num_columnas > 5:
            font_size_header = 9
            font_size_body = 8
        else:
            font_size_header = 10
            font_size_body = 9

        # Estilos para celdas con word wrap
        header_cell_style = ParagraphStyle(
            'HeaderCell',
            parent=styles['Normal'],
            fontSize=font_size_header,
            fontName='Helvetica-Bold',
            textColor=colors.white,
            alignment=TA_CENTER,
            wordWrap='CJK',
            leading=font_size_header + 2
        )

        body_cell_style = ParagraphStyle(
            'BodyCell',
            parent=styles['Normal'],
            fontSize=font_size_body,
            fontName='Helvetica',
            alignment=TA_LEFT,
            wordWrap='CJK',
            leading=font_size_body + 2
        )

        body_cell_right_style = ParagraphStyle(
            'BodyCellRight',
            parent=body_cell_style,
            alignment=TA_RIGHT
        )

        body_cell_center_style = ParagraphStyle(
            'BodyCellCenter',
            parent=body_cell_style,
            alignment=TA_CENTER
        )

        totals_cell_style = ParagraphStyle(
            'TotalsCell',
            parent=body_cell_style,
            fontName='Helvetica-Bold'
        )

        totals_cell_right_style = ParagraphStyle(
            'TotalsCellRight',
            parent=totals_cell_style,
            alignment=TA_RIGHT
        )

        # Título
        elements.append(Paragraph(f"Informe de {nombre_reporte}", title_style))

        # Fecha de generación
        fecha_generacion = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elements.append(Paragraph(f"Generado: {fecha_generacion}", date_style))

        # Todas las columnas visibles
        columnas_visibles = columnas

        table_data = []

        # Encabezados - usar Paragraph para permitir wrap
        headers = []
        for col in columnas_visibles:
            header_text = col['nombre_mostrar']
            headers.append(Paragraph(header_text, header_cell_style))
        table_data.append(headers)

        # Datos - usar Paragraph para permitir wrap
        for fila in datos:
            row = []
            for col in columnas_visibles:
                valor = fila.get(col['campo'], '')
                valor_formateado = self._formatear_valor_pdf(
                    valor,
                    col.get('tipo_dato', 'string'),
                    col.get('decimales', 2),
                    col.get('sufijo', ''),
                    col.get('campo', '')
                )

                # Seleccionar estilo según alineación de la columna
                alineacion = col.get('alineacion', 'left')
                if alineacion == 'right':
                    cell_style = body_cell_right_style
                elif alineacion == 'center':
                    cell_style = body_cell_center_style
                else:
                    cell_style = body_cell_style

                row.append(Paragraph(str(valor_formateado), cell_style))
            table_data.append(row)

        # Fila de totales
        if totales:
            total_row = [Paragraph('TOTALES', totals_cell_style)]
            for col in columnas_visibles[1:]:
                campo = col['campo']
                if campo in totales:
                    valor = totales[campo]
                    valor_formateado = self._formatear_valor_pdf(
                        valor,
                        col.get('tipo_dato', 'number'),
                        col.get('decimales', 2),
                        col.get('sufijo', ''),
                        campo
                    )
                    # Totales numéricos alineados a la derecha
                    alineacion = col.get('alineacion', 'right')
                    if alineacion == 'right':
                        total_row.append(Paragraph(str(valor_formateado), totals_cell_right_style))
                    else:
                        total_row.append(Paragraph(str(valor_formateado), totals_cell_style))
                else:
                    total_row.append(Paragraph('', totals_cell_style))
            table_data.append(total_row)

        # Calcular anchos de columna dinámicamente
        available_width = pagesize[0] - (2 * margins)
        col_widths = self._calcular_anchos_columnas_pdf(
            columnas_visibles,
            datos,
            available_width
        )

        # Crear tabla con repeatRows para repetir encabezado en cada página
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Estilos de tabla
        style = TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{self.HEADER_COLOR}')),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),

            # Cuerpo - alinear arriba para celdas con wrap
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),

            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),

            # Línea gruesa debajo del encabezado
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor(f'#{self.HEADER_COLOR}')),
        ])

        # Filas alternadas
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor(f'#{self.ALTERNATE_ROW_COLOR}'))

        # Fila de totales
        if totales:
            last_row = len(table_data) - 1
            style.add('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor(f'#{self.TOTALS_COLOR}'))

        table.setStyle(style)
        elements.append(table)

        # Pie de página con total de registros
        elements.append(Spacer(1, 12))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.gray
        )
        elements.append(Paragraph(f"Total de registros: {len(datos)}", footer_style))

        # Construir PDF
        doc.build(elements)

        buffer.seek(0)
        return buffer.getvalue()

    def _calcular_anchos_columnas_pdf(
            self,
            columnas: List[Dict[str, Any]],
            datos: List[Dict[str, Any]],
            page_width: float
    ) -> List[float]:
        """
        Calcula anchos de columna basados en el tipo de contenido.
        Asigna más espacio a columnas de texto largo (material, descripción, nombre).

        Args:
            columnas: Definición de columnas
            datos: Datos del reporte
            page_width: Ancho disponible de la página

        Returns:
            Lista de anchos de columna
        """
        # Calcular pesos según tipo de columna
        pesos = []
        for col in columnas:
            nombre = col.get('nombre_mostrar', col.get('campo', '')).lower()
            tipo_dato = col.get('tipo_dato', 'string')

            # Columnas de texto largo: más peso
            if any(x in nombre for x in ['material', 'descripcion', 'descripción', 'nombre', 'observacion', 'observación', 'almacenamiento', 'detalle']):
                pesos.append(3.0)
            # Columnas de fecha
            elif tipo_dato in ('date', 'datetime') or 'fecha' in nombre:
                pesos.append(1.5)
            # Columnas numéricas
            elif tipo_dato == 'number' or any(x in nombre for x in ['cantidad', 'saldo', 'total', 'peso', 'valor', 'precio']):
                pesos.append(2.0)
            # ID y campos cortos
            elif any(x in nombre for x in ['id', 'código', 'codigo', 'consecutivo']):
                pesos.append(1.0)
            else:
                pesos.append(1.5)

        # Calcular anchos proporcionales
        total_pesos = sum(pesos)
        anchos = [(p / total_pesos) * page_width for p in pesos]

        # Aplicar límites mínimos y máximos
        min_width = 1.5 * cm
        max_width = page_width * 0.4  # Máximo 40% del ancho para una columna

        anchos_ajustados = []
        for ancho in anchos:
            anchos_ajustados.append(max(min_width, min(max_width, ancho)))

        # Ajustar para que quepan en la página
        total_width = sum(anchos_ajustados)
        if total_width != page_width:
            scale = page_width / total_width
            anchos_ajustados = [a * scale for a in anchos_ajustados]

        return anchos_ajustados

    def _formatear_valor_pdf(
            self,
            valor: Any,
            tipo_dato: str,
            decimales: int = 2,
            sufijo: str = '',
            campo: str = ''
    ) -> str:
        """
        Formatea un valor para PDF.

        Args:
            valor: Valor a formatear
            tipo_dato: Tipo de dato de la columna
            decimales: Número de decimales para números
            sufijo: Sufijo a agregar al valor (ej: unidad de medida)
            campo: Nombre del campo (para detectar campos enteros)
        """
        if valor is None:
            return '-'

        # Verificar si es un campo que debe ser entero
        campo_lower = campo.lower() if campo else ''
        es_campo_entero = campo_lower in self.CAMPOS_ENTEROS or tipo_dato == 'integer'


        if es_campo_entero:
            try:
                numero = int(float(valor))
                formateado = f"{numero:,}"
                if sufijo:
                    formateado = f"{formateado} {sufijo}"
                return formateado
            except (ValueError, TypeError):
                return str(valor)

        if tipo_dato == 'number':
            try:
                numero = float(valor)
                formateado = f"{numero:,.{decimales}f}"
                if sufijo:
                    formateado = f"{formateado} {sufijo}"
                return formateado
            except (ValueError, TypeError):
                return str(valor)

        elif tipo_dato == 'date':
            if isinstance(valor, datetime):
                return valor.strftime('%Y-%m-%d')
            elif isinstance(valor, str):
                try:
                    dt = datetime.fromisoformat(valor.replace('Z', '+00:00'))
                    return dt.strftime('%Y-%m-%d')
                except:
                    return valor

        elif tipo_dato == 'datetime':
            if isinstance(valor, datetime):
                return valor.strftime('%Y-%m-%d %H:%M')
            elif isinstance(valor, str):
                try:
                    dt = datetime.fromisoformat(valor.replace('Z', '+00:00'))
                    return dt.strftime('%Y-%m-%d %H:%M')
                except:
                    return valor

        return str(valor) if valor else '-'

    # ========================================================
    # EXPORTACIÓN A CSV
    # ========================================================

    def exportar_csv(
            self,
            datos: List[Dict[str, Any]],
            columnas: List[Dict[str, Any]]
    ) -> str:
        """
        Exporta datos a formato CSV.

        Args:
            datos: Lista de registros a exportar
            columnas: Definición de columnas

        Returns:
            Contenido del archivo CSV como string
        """
        # Crear DataFrame con las columnas en el orden correcto
        campos = [col['campo'] for col in columnas]
        nombres = {col['campo']: col['nombre_mostrar'] for col in columnas}

        # Extraer solo los campos definidos
        datos_filtrados = []
        for fila in datos:
            fila_filtrada = {campo: fila.get(campo, '') for campo in campos}
            datos_filtrados.append(fila_filtrada)

        df = pd.DataFrame(datos_filtrados)

        # Reordenar columnas
        df = df[campos]

        # Renombrar columnas
        df.rename(columns=nombres, inplace=True)

        # Formatear valores
        for col in columnas:
            campo = col['campo']
            nombre = col['nombre_mostrar']
            tipo = col.get('tipo_dato', 'string')

            if nombre in df.columns:
                # Verificar si es un campo que debe ser entero
                campo_lower = campo.lower()
                es_campo_entero = campo_lower in self.CAMPOS_ENTEROS or tipo == 'integer'


                if es_campo_entero:
                    df[nombre] = df[nombre].apply(
                        lambda x: self._formatear_entero_csv(x)
                    )
                elif tipo == 'datetime':
                    df[nombre] = df[nombre].apply(
                        lambda x: self._formatear_datetime_csv(x)
                    )
                elif tipo == 'date':
                    df[nombre] = df[nombre].apply(
                        lambda x: self._formatear_date_csv(x)
                    )

        # Convertir a CSV
        return df.to_csv(index=False, encoding='utf-8')

    def _formatear_datetime_csv(self, valor: Any) -> str:
        """Formatea datetime para CSV."""
        if valor is None:
            return ''
        if isinstance(valor, datetime):
            return valor.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(valor, str):
            try:
                dt = datetime.fromisoformat(valor.replace('Z', '+00:00'))
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                return valor
        return str(valor) if valor else ''

    def _formatear_date_csv(self, valor: Any) -> str:
        """Formatea date para CSV."""
        if valor is None:
            return ''
        if isinstance(valor, datetime):
            return valor.strftime('%Y-%m-%d')
        elif isinstance(valor, str):
            try:
                dt = datetime.fromisoformat(valor.replace('Z', '+00:00'))
                return dt.strftime('%Y-%m-%d')
            except:
                return valor
        return str(valor) if valor else ''

    def _formatear_entero_csv(self, valor: Any) -> str:
        """
        Formatea un valor como entero para CSV (sin decimales).

        Args:
            valor: Valor a formatear

        Returns:
            Valor formateado como entero (string)
        """
        if valor is None:
            return ''
        try:
            return str(int(float(valor)))
        except (ValueError, TypeError):
            return str(valor) if valor else ''
